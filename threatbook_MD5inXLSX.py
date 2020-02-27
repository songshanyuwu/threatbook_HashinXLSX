import requests
import json
import openpyxl
import time


def detectMD5(md5, workSheet, rowNum):
    # 方法用途，获取病毒木马的MD5、worksheet实例、所遍历到的行数进度rowNum，通过调用微步api查到威胁情报，并将相关内容写入worksheet实例相应的单元格中
    url_summary = 'https://s.threatbook.cn/api/v2/file/report/summary'
    url_multiengines = 'https://s.threatbook.cn/api/v3/file/report/multiengines'
    params = {
        # 微步个人中心的apikey放在下面
        'apikey': 'apikey放在这里',
        'md5': md5
    }
    print('第'+str(rowNum-1)+'个病毒木马检测，MD5：\n'+md5)
    # 请求api：report/summary
    response_summary = requests.get(url_summary, params=params)
    rejson_summary = json.loads(json.dumps(response_summary.json()))
    # 请求api：report/multiengines
    response_multiengines = requests.get(url_multiengines, params=params)
    rejson_multiengines = json.loads(json.dumps(response_multiengines.json()))
    if(rejson_summary['response_code'] != 0):
        print('暂无结果：', rejson_summary['msg'])
        workSheet.cell(row=rowNum, column=2).value = rejson_summary['msg']
    else:
        workSheet.cell(row=rowNum, column=2).value = '查询成功'
        level = rejson_summary['data']['summary']['threat_level']
        levelMeans = ''
        print('【', level, '】')
        if (level == 'malicious'):
            levelMeans = '恶意'
        elif (level == 'suspicious'):
            levelMeans = '可疑'
        elif (level == 'clean'):
            levelMeans = '安全'
        elif (level == ''):
            levelMeans = '无建议'
        print('威胁等级:', levelMeans)
        workSheet.cell(
            row=rowNum, column=3).value = levelMeans
        print('检测标签为：', rejson_summary['data']['summary']['tag']['x'])
        workSheet.cell(
            row=rowNum, column=4).value = str(rejson_summary['data']['summary']['tag']['x'])
        print('反病毒扫描引擎检出率:',
              rejson_summary['data']['summary']['multi_engines'])
        workSheet.cell(
            row=rowNum, column=5).value = rejson_summary['data']['summary']['multi_engines']
        if(rejson_multiengines['response_code'] == 0):
            print('病毒家族：', rejson_multiengines['data']
                  ['multiengines']['malware_family'])
            workSheet.cell(
                row=rowNum, column=6).value = str(rejson_multiengines['data']['multiengines']['malware_family'])


def loadMD5(xlsPath):
    wb = openpyxl.load_workbook(xlsPath)
    ws = wb.get_sheet_by_name('Sheet1')
    ws.cell(row=1, column=2).value = '微步查询状态'
    ws.cell(row=1, column=3).value = '威胁等级建议'
    ws.cell(row=1, column=4).value = '检测标签'
    ws.cell(row=1, column=5).value = '反病毒扫描引擎检出率'
    ws.cell(row=1, column=6).value = '病毒家族'
    wb.save(xlsPath)
    # 获取表格的最大行数
    cols = ws.max_row
    # 起始位置i=2跳过标题行
    i = 2
    while i <= cols:
        # 判断当‘微步查询状态’为空，或上次查询状态为‘IN_PROGRESS’的会执行查询
        if(ws.cell(row=i, column=2).value is None or ws.cell(row=i, column=2).value == 'IN_PROGRESS'):
            detectMD5(ws.cell(row=i, column=1).value, ws, i)
            wb.save(xlsPath)
            # 免费api频率限制为每分钟不超过100次，sleep 0.6秒保证不超
            time.sleep(0.6)
        i += 1


if __name__ == '__main__':
    loadMD5('d:/MD5.xlsx')
